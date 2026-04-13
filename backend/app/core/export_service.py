"""Data export and report generation service."""
import io
import csv
from datetime import datetime
from typing import Optional, List, Dict, Any
from pathlib import Path

import pandas as pd
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.student import Student


class ExportService:
    """Service for exporting student data to various formats."""
    
    # Displayable field names
    FIELD_NAMES = {
        'id': 'ID',
        'first_name': 'First Name',
        'last_name': 'Last Name',
        'email': 'Email',
        'class_name': 'Class',
        'roll_no': 'Roll No.',
        'admission_no': 'Admission No.',
        'mobile_numbers': 'Mobile Numbers',
        'address': 'Address',
        'city': 'City',
        'state': 'State',
        'zip_code': 'Zip Code',
        'date_of_birth': 'Date of Birth',
        'enrollment_date': 'Enrollment Date',
        'is_active': 'Status',
        'notes': 'Notes',
    }
    
    @staticmethod
    def _serialize_student(student: Student, fields: List[str]) -> Dict[str, Any]:
        """Convert student object to dictionary with selected fields."""
        data = {}
        for field in fields:
            if field not in ExportService.FIELD_NAMES:
                continue
            
            value = getattr(student, field, None)
            
            # Handle special formatting
            if field == 'is_active':
                value = 'Active' if value else 'Inactive'
            elif field == 'mobile_numbers' and isinstance(value, list):
                value = '; '.join(value)
            elif isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d')
            
            data[ExportService.FIELD_NAMES[field]] = value
        
        return data
    
    @staticmethod
    def export_to_csv(
        students: List[Student],
        fields: List[str],
        filename: Optional[str] = None,
    ) -> io.BytesIO:
        """Export students to CSV format."""
        if not filename:
            filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        output = io.StringIO()
        
        if not students or not fields:
            return io.BytesIO(output.getvalue().encode('utf-8'))
        
        # Get field headers
        headers = [ExportService.FIELD_NAMES.get(f, f) for f in fields]
        
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        
        for student in students:
            row = ExportService._serialize_student(student, fields)
            writer.writerow(row)
        
        # Convert to BytesIO
        csv_bytes = output.getvalue().encode('utf-8')
        return io.BytesIO(csv_bytes), filename
    
    @staticmethod
    def export_to_excel(
        students: List[Student],
        fields: List[str],
        title: str = "Student Report",
        group_by: Optional[str] = None,
        filename: Optional[str] = None,
    ) -> tuple[io.BytesIO, str]:
        """Export students to Excel format with formatting."""
        if not filename:
            filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Add title
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:Z1')
        
        # Add export date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:Z2')
        
        # Add headers
        headers = [ExportService.FIELD_NAMES.get(f, f) for f in fields]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data rows
        if group_by and group_by in fields:
            # Group data
            grouped = {}
            for student in students:
                group_value = getattr(student, group_by, 'Other')
                if group_value not in grouped:
                    grouped[group_value] = []
                grouped[group_value].append(student)
            
            row = 5
            for group_name, group_students in sorted(grouped.items()):
                # Add group header
                ws.insert_rows(1, 1)
                group_cell = ws.cell(row=row, column=1)
                group_cell.value = f"{ExportService.FIELD_NAMES.get(group_by, group_by)}: {group_name}"
                group_cell.font = Font(bold=True, size=11, color="FFFFFF")
                group_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                ws.merge_cells(f'A{row}:Z{row}')
                row += 1
                
                # Add group data
                for student in group_students:
                    data = ExportService._serialize_student(student, fields)
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = data.get(header)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'),
                        )
                    row += 1
        else:
            # No grouping
            for row_idx, student in enumerate(students, 5):
                data = ExportService._serialize_student(student, fields)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = data.get(header)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'),
                    )
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Save to BytesIO
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        return excel_bytes, filename
    
    @staticmethod
    def export_to_pdf(
        students: List[Student],
        fields: List[str],
        title: str = "Student Report",
        group_by: Optional[str] = None,
        filename: Optional[str] = None,
    ) -> tuple[io.BytesIO, str]:
        """Export students to PDF format."""
        if not filename:
            filename = f"students_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        pdf_bytes = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_bytes,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12,
            alignment=TA_CENTER,
        )
        story.append(Paragraph(title, title_style))
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6b7280'),
            spaceAfter=12,
            alignment=TA_CENTER,
        )
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style))
        
        # Prepare table data
        headers = [ExportService.FIELD_NAMES.get(f, f) for f in fields]
        table_data = [headers]
        
        for student in students:
            row = []
            data = ExportService._serialize_student(student, fields)
            for header in headers:
                row.append(str(data.get(header, '')))
            table_data.append(row)
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        pdf_bytes.seek(0)
        
        return pdf_bytes, filename
    
    @staticmethod
    def export_students(
        students: List[Student],
        fields: List[str],
        export_format: str = "csv",
        title: str = "Student Report",
        group_by: Optional[str] = None,
        filename: Optional[str] = None,
    ) -> tuple[io.BytesIO, str]:
        """
        Export students to the specified format.
        
        Args:
            students: List of Student objects
            fields: List of field names to include
            export_format: 'csv', 'excel', or 'pdf'
            title: Report title
            group_by: Field to group by (for Excel/PDF)
            filename: Output filename
        
        Returns:
            Tuple of (BytesIO object, filename)
        """
        if export_format == "csv":
            return ExportService.export_to_csv(students, fields, filename)
        elif export_format == "excel":
            return ExportService.export_to_excel(students, fields, title, group_by, filename)
        elif export_format == "pdf":
            return ExportService.export_to_pdf(students, fields, title, group_by, filename)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
